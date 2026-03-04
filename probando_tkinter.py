import tkinter as tk
from tkinter import END, Text, Spinbox, Toplevel, Scale, HORIZONTAL  # import *

#1
root = tk.Tk()
root.title("Saludos")
#root.geometry("400x300")
#2
tk.Label(root, text="Nombre").grid(row=0, column=0) # o etiqueta.pack()
tk.Label(root, text="Apellidos").grid(row=1, column=0)
tk.Entry(root).grid(row=0, column=1)
tk.Entry(root).grid(row=1, column=1)
#3
var1 = tk.IntVar()#1 o 0
var2 = tk.IntVar()
tk.Checkbutton(root, text="Mesi", variable=var1).grid(row=3, sticky="w")# o sticky=tk.W o si es sin grid: pack(side=tk.LEFT)
tk.Checkbutton(root, text="Cr7", variable=var2).grid(row=4, sticky="w")
#4
v = tk.IntVar()
tk.Radiobutton(root, text="A", variable=v, value=1).grid(row=5, sticky="w")
tk.Radiobutton(root, text="B", variable=v, value=2).grid(row=6, sticky="w")
#5
lb = tk.Listbox(root)
lb.insert(1, "Python")
lb.insert(2, "Java")
lb.insert(3, "C++")
lb.insert(4, "Any other")
lb.grid(row=7, column=0, columnspan=2)
#6
text_widget = Text(root, height=2, width=30)
text_widget.insert(END, "Frase de varias\nlineas jeje\n")
text_widget.grid(row=9, column=0)
#7
spinbox = Spinbox(root, from_=0, to=10)
spinbox.grid(row=10, column=0)
#8
scrollbar = tk.Scrollbar(root)
scrollbar.grid(row=8, column=0)

mylist = tk.Listbox(root, yscrollcommand=scrollbar.set)#-> connects the listbox to the scrollbar.

for line in range(100):
    mylist.insert(tk.END, "This is line number " + str(line))

mylist.grid(row=8, column=0, columnspan=1)
scrollbar.config(command=mylist.yview)# Enables scrolling behavior
#9
menu = tk.Menu(root)
root.config(menu=menu)

filemenu = tk.Menu(menu)
menu.add_cascade(label="File", menu=filemenu)#-> adds dropdown menus
filemenu.add_command(label="New")
filemenu.add_command(label="Open...")
filemenu.add_separator()
filemenu.add_command(label="Exit", command=root.quit)

helpmenu = tk.Menu(menu)
menu.add_cascade(label="Help", menu=helpmenu)
helpmenu.add_command(label="About")
#10
top = Toplevel()# (abrir ventana vacia)
top.title('Python')
#11
vertical_scale = Scale(root, from_=0, to=42)
vertical_scale.grid(row=11, column=0)
horizontal_scale = Scale(root, from_=0, to=200, orient="horizontal")
horizontal_scale.grid(row=11, column=1)
#12 y 13 -> Faltan MenuButton y Progressbar

#14
boton = tk.Button(root, text="Cerrar", width=25, command=root.destroy)
boton.grid(row=12, column=1, columnspan=2, pady=10)

root.mainloop()


"""
import openpyxl
import requests
from bs4 import BeautifulSoup

BASE_URL = "https://store.steampowered.com/search/?page="

headers = {
    "User-Agent": "Mozilla/5.0"
}

def guardar_en_excel(fila):
    archivo = "JuegosExcelBD.xlsx"
    try:
        wb = openpyxl.load_workbook(archivo)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"]) #hoja por defecto
        sheet = wb.create_sheet("Juegos")
        sheet.append(["id", "titulo", "precio", "precio anterior", "descuento", "released", "imagen", "so", "fecha_scrap"])
    else:
        sheet = wb["Juegos"]

    sheet.append(fila)
    wb.save(archivo)

def sopear(url, page):
    response = requests.get(url, headers=headers, verify=False)

    print(f"PÁGINA {page} → STATUS: {response.status_code}")

    if response.status_code != 200:
        print("Bloqueo detectado.")
        return None

    return BeautifulSoup(response.content, "html.parser")

arrayJuegos = []
dictJuego = {}
for page in range(1, 3):
    url = BASE_URL + str(page)
    html = sopear(url, page)

    if html:
        juegos = html.find_all("a", class_="search_result_row")

        for juego in juegos:
            dictJuego = {}
            id_steam = juego.get("data-ds-appid")
            titulo = juego.find("span", class_="title").text
            precio = juego.find("div", class_="discount_final_price").text if juego.find("div", class_="discount_final_price").text != "Free" else "0,00€"
            imagen = juego.find("img").get("src")
            so_nativos = [p.get("class")[1] for p in juego.find_all("span", class_="platform_img") # <span class="platform_img win"></span>
                          if p.get("class")[1] != "group_separator"]  # <span class="platform_img group_separator"></span> -> no guardar
            released = juego.find("div", class_="search_released").text.strip()
            discount_original_price = juego.find("div", class_="discount_original_price").text.strip() if juego.find("div", class_="discount_original_price") else None
            discount_pct = juego.find("div", class_="discount_pct").text if juego.find("div", class_="discount_pct") else None

            dictJuego["id steam"] = id_steam
            es_vr = True if juego.find("span", class_="vr_supported") else False
            dictJuego["titulo"] = titulo if not es_vr else titulo + "(VR Supported)"
            dictJuego["precio"] = precio[:-1] # quitar '€'
            if discount_original_price:
                dictJuego["precio anterior"] = discount_original_price[:-1]
            if discount_pct:
                dictJuego["porcentaje descuento"] = discount_pct[1:-1]  # -90%
            dictJuego["imagen"] = imagen
            dictJuego["sistemas operativos"] = so_nativos

            dictJuego["released"] = released
            dictJuego["fecha_scrap"] = "2024-03-04"  # fecha actual

            arrayJuegos.append(dictJuego)
            print(dictJuego)

"""